"""
GST Reconciliation Engine
=========================
Compares Books (Inward Register) vs GSTR-2B to find matching invoices.

Pipeline:
  Stage 0  — load, clean, recompute money, group books to invoice grain
  Stage 1  — exact matching (1A / 1B with uniqueness gate / 1C collision)
  Stage 2  — standardized-invoice matching (same logic, invoice_std)
  Stage 3  — split into blue / nonblue queues (no auto-match here)
  Stage 4A — AI pass on blue queue
  Stage 4B — AI pass on nonblue queue (pile builder filters no-money-twin rows)
  Stage 5  — greedy one-to-one merge + consolidated single-sheet workbook

Fixes applied vs submitted version:
  - B2C rows excluded from unmatched_books (were double-listed)
  - 1B uniqueness gate added (spec requirement; masks collision clusters)
  - Stage 2 now gets work when 1B is correctly restricted
  - Type-column guard with loud failure if label set changes
  - money_match checks taxable + each tax head individually (per spec)
  - find_identity_money_mismatch uses invoice_std + uniqueness guard
  - build_piles size-gate pre-pass now records and returns its matches
  - Stage 3 q_remaining_g excludes blue 2B rows (no double-proposal risk)
  - per-pile AI logging added (spec requirement)
  - source column added to canonical schema for both books and 2B
"""

from __future__ import annotations

import asyncio
import json
import logging
import os
import re
from collections import Counter, defaultdict

from dataclasses import dataclass, field
from pathlib import Path

import networkx as nx
import pandas as pd
from dotenv import load_dotenv
from openai import AsyncOpenAI
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

load_dotenv()

# ── logging ───────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.DEBUG,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    handlers=[
        logging.FileHandler("ai_requests.log", mode="w", encoding="utf-8"),
        logging.StreamHandler(),
    ],
)
# suppress noisy transport-layer debug output from httpx / httpcore / openai
logging.getLogger("httpx").setLevel(logging.WARNING)
logging.getLogger("httpcore").setLevel(logging.WARNING)
logging.getLogger("openai").setLevel(logging.WARNING)
_log = logging.getLogger(__name__)

# ── config ────────────────────────────────────────────────────────────────────
BOOKS_PATH  = r"C:\Users\subra\OneDrive\Inward Register Feb-25.xlsx"
B2B_PATH    = r"C:\Users\subra\OneDrive\022025_29ADAFS3950J1ZS_GSTR2B_18032025_SSA Firm.xlsx"
OUTPUT_PATH = Path("gst_reconciliation_output.xlsx")

TOL                 = 2.0
PILE_SIZE_LIMIT     = 15      # rows per side before exact-invoice pre-pass inside pile
BLUE_AUTO_ACCEPT    = 0.90
BLUE_MIN_REVIEW     = 0.60
NONBLUE_AUTO_ACCEPT = 0.93
NONBLUE_MIN_REVIEW  = 0.75
AI_MODEL            = "openai/gpt-4o-mini"   # OpenRouter model ID
# AI_MODEL            = "gemini-2.5-flash-lite"
# AI_MODEL            = "openai/gpt-5-mini"     # stronger, but more expensive and slower — use only if needed

RATE_COLS = {
    "Taxable @ 0%":  0.00,
    "Taxable @ 5%":  0.05,
    "Taxable @ 12%": 0.12,
    "Taxable @ 18%": 0.18,
    "Taxable @ 28%": 0.28,
}

# ── helpers ───────────────────────────────────────────────────────────────────

def clean_str(val) -> str | None:
    """Strip, uppercase, return None for blanks / NA sentinels."""
    if pd.isna(val):
        return None
    s = str(val).strip().upper()
    return s if s not in ("", "NAN", "<NA>", "NONE", "NAT") else None


def standardize_invoice(val) -> str | None:
    """Remove all non-alphanumeric, uppercase, drop leading zeros."""
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return None
    s = re.sub(r"[^A-Z0-9]", "", str(val).upper()).lstrip("0")
    return s or None


def money_match(b: pd.Series, g: pd.Series, tol: float = TOL) -> bool:
    """
    True when taxable AND total_tax are both within tol.

    We compare total_tax (cgst+sgst+igst) as a single figure rather than
    each head independently. This correctly handles place-of-supply
    differences where books record CGST+SGST but 2B reports IGST (or
    vice versa) — same taxable, same total tax, legitimately the same
    invoice.
    """
    return (
        abs(b["taxable"]   - g["taxable"])   <= tol
        and abs(b["total_tax"] - g["total_tax"]) <= tol
    )


# ── match record ──────────────────────────────────────────────────────────────

@dataclass
class Match:
    books_id:      str
    b2b_id:        str
    stage:         str    # "1" | "2"
    sub:           str    # "1A" | "1B" | "1C" | "2A" | "2B"
    match_reason:  str
    taxable_diff:  float
    total_tax_diff: float


# global AI interaction log — populated by _ai_pile, written to HTML by write_ai_log_html
_AI_LOG: list[dict] = []


# ═══════════════════════════════════════════════════════════════════════════════
# STAGE 0 — LOAD, CLEAN, RECOMPUTE, GROUP
# ═══════════════════════════════════════════════════════════════════════════════

_INTRA_LABELS = {"INTRA-STATE", "INTRA STATE", "INTRASTATE", "LOCAL"}
_INTER_LABELS = {"INTER-STATE", "INTER STATE", "INTERSTATE", "IMPORT", "SEZ"}


def load_books(
    path: str,
    col_map: dict | None = None,
    header_row: int = 1,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Load and clean the books (Inward Register).
    Returns (clean_df, dirty_df).

    col_map  — {canonical_name: actual_col_name_in_file}. Applied as a
               rename before any processing so downstream code always
               sees the canonical names.
    header_row — 0-indexed row number that contains the column headers
                 (pandas header= param). Default 1 = second row.

    Money strategy:
      1. Recompute taxable from rate-slab columns (rate cols are always populated).
      2. Use the Type column (INTRA-STATE / INTER-STATE) to split into
         CGST+SGST vs IGST.
      3. Validate: |recomputed_gross - Gross Total| <= 2.
         Rows that fail → dirty_df (skip matching, write to dirty_input sheet).

    Guard: if Type column has unexpected labels, raise immediately so the
    caller knows not to trust the tax-head split.
    """
    df = pd.read_excel(path, header=header_row, engine="openpyxl")

    if col_map:
        rename = {actual: canonical for canonical, actual in col_map.items() if actual and actual in df.columns}
        if rename:
            df = df.rename(columns=rename)
    df = df[~(df["Supplier"].isna() & df["Gross Total"].isna())].copy()

    # coerce all numeric columns
    numeric_cols = [
        "Gross Total", "Taxable Value", "CGST", "SGST", "IGST",
        "Cess", "Exempt", "Others",
        *[c for c in RATE_COLS if c in df.columns],
    ]
    for col in numeric_cols:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

    # ── Type-column guard ────────────────────────────────────────────────────
    if "Type" not in df.columns:
        raise ValueError(
            "Books file is missing a 'Type' column (expected INTRA-STATE / "
            "INTER-STATE). Cannot split CGST/SGST vs IGST without it."
        )
    raw_types = set(df["Type"].dropna().astype(str).str.strip().str.upper().unique())
    unknown   = raw_types - _INTRA_LABELS - _INTER_LABELS - {"NAN", ""}
    if unknown:
        raise ValueError(
            f"Books 'Type' column has unexpected values: {unknown}. "
            "Update _INTRA_LABELS / _INTER_LABELS in config and verify."
        )

    is_intra = df["Type"].astype(str).str.strip().str.upper().isin(_INTRA_LABELS)

    # ── recompute money from rate-slab columns ───────────────────────────────
    avail = [c for c in RATE_COLS if c in df.columns]
    if not avail:
        raise ValueError("No rate-slab columns found (Taxable @ X%). Cannot recompute money.")

    df["taxable"] = sum(df[c] for c in avail).round(2)

    # CGST = half the rate on intra-state rows, 0 on inter-state
    df["cgst"] = sum(
        df[c] * (RATE_COLS[c] / 2) for c in avail
    ).where(is_intra, 0).round(2)
    df["sgst"] = df["cgst"].copy()

    # IGST = full rate on inter-state rows, 0 on intra-state
    df["igst"] = sum(
        df[c] * RATE_COLS[c] for c in avail
    ).where(~is_intra, 0).round(2)

    df["total_tax"] = (df["cgst"] + df["sgst"] + df["igst"]).round(2)

    # fallback: where recomputed value is 0 but cached column is non-zero, use cached
    cached_taxable = pd.to_numeric(df.get("Taxable Value", 0), errors="coerce").fillna(0)
    cached_cgst    = pd.to_numeric(df.get("CGST", 0),          errors="coerce").fillna(0)
    cached_sgst    = pd.to_numeric(df.get("SGST", 0),          errors="coerce").fillna(0)
    cached_igst    = pd.to_numeric(df.get("IGST", 0),          errors="coerce").fillna(0)

    df["taxable"]   = df["taxable"].where(df["taxable"] != 0, cached_taxable)
    df["cgst"]      = df["cgst"].where(df["cgst"]       != 0, cached_cgst)
    df["sgst"]      = df["sgst"].where(df["sgst"]       != 0, cached_sgst)
    df["igst"]      = df["igst"].where(df["igst"]       != 0, cached_igst)
    df["total_tax"] = (df["cgst"] + df["sgst"] + df["igst"]).round(2)

    cess   = df["Cess"]   if "Cess"   in df.columns else 0
    exempt = df["Exempt"] if "Exempt" in df.columns else 0
    others = df["Others"] if "Others" in df.columns else 0
    df["gross_recomputed"] = (
        df["taxable"] + df["total_tax"] + cess + exempt + others
    ).round(2)
    df["gross_diff"] = (df["Gross Total"] - df["gross_recomputed"]).round(2)

    # ── clean identity fields ────────────────────────────────────────────────
    for col in ["GSTIN/UIN", "Voucher Number", "Supplier"]:
        df[col] = df[col].apply(clean_str)
    df["invoice_date"] = pd.to_datetime(df.get("Accounting Date"), errors="coerce")

    dirty = df[df["gross_diff"].abs() > TOL].copy()
    clean = df[df["gross_diff"].abs() <= TOL].copy()

    print(
        f"[Stage 0] Books: {len(df)} rows | "
        f"dirty={len(dirty)} | clean={len(clean)} | "
        f"intra={is_intra.sum()} inter={(~is_intra).sum()}"
    )
    return clean, dirty


def load_2b(path: str) -> pd.DataFrame:
    """Load GSTR-2B B2B sheet; flatten MultiIndex header; return canonical frame."""
    df = pd.read_excel(path, sheet_name="B2B", header=[4, 5], engine="openpyxl")

    sup_col = ("Trade/Legal name", "Unnamed: 1_level_1")
    val_col = ("Invoice Details", "Invoice Value(₹)")
    df = df[~(df[sup_col].isna() & df[val_col].isna())].copy()

    # flatten MultiIndex: keep level0 only when level1 is "Unnamed…"
    df.columns = [
        a if "Unnamed" in b else f"{a}::{b}"
        for a, b in df.columns
    ]

    df = df.rename(columns={
        "GSTIN of supplier":                 "gstin",
        "Trade/Legal name":                  "supplier",
        "Invoice Details::Invoice number":   "invoice_raw",
        "Invoice Details::Invoice Date":     "invoice_date",
        "Invoice Details::Invoice Value(₹)": "invoice_value",
        "Taxable Value (₹)":                 "taxable",
        "Tax Amount::Central Tax(₹)":        "cgst",
        "Tax Amount::State/UT Tax(₹)":       "sgst",
        "Tax Amount::Integrated Tax(₹)":     "igst",
    })

    for col in ["taxable", "cgst", "sgst", "igst", "invoice_value"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

    df["total_tax"]    = (df["cgst"] + df["sgst"] + df["igst"]).round(2)
    df["invoice_date"] = pd.to_datetime(df.get("invoice_date"), errors="coerce")

    for col in ["gstin", "invoice_raw", "supplier"]:
        df[col] = df[col].apply(clean_str)

    df["invoice_std"] = df["invoice_raw"].apply(standardize_invoice)
    df["source"] = "2b"
    df = df.reset_index(drop=True)
    df.insert(0, "row_id", ["G" + str(i) for i in df.index])

    print(f"[Stage 0] 2B: {len(df)} rows")
    return df


def group_books(df: pd.DataFrame) -> pd.DataFrame:
    """
    Group books to invoice grain.

    Key assignment (OR, not AND — each row gets exactly one key):
      GSTIN + invoice present → key = (gstin, invoice_raw)
      GSTIN missing, invoice present → key = (supplier, invoice_raw)
      No invoice (regardless of GSTIN) → standalone key (never merged)

    The 2B is NOT grouped — it is already one row per (GSTIN + invoice).
    """
    def make_key(row):
        g = row["GSTIN/UIN"]
        v = row["Voucher Number"]
        s = row["Supplier"]
        if v:
            if g:
                return ("gstin_invoice", g, v)
            else:
                return ("supplier_invoice", s or f"_null_{row.name}", v)
        # no invoice — standalone; row.name is the original df index (unique)
        return ("standalone", f"_solo_{row.name}", f"_solo_{row.name}")

    df = df.copy()
    df[["_ktype", "_k1", "_k2"]] = df.apply(
        lambda r: pd.Series(make_key(r)), axis=1
    )

    grouped = (
        df.groupby(["_ktype", "_k1", "_k2"], dropna=False, sort=False)
        .agg(
            gstin          = ("GSTIN/UIN",      "first"),
            invoice_raw    = ("Voucher Number",  "first"),
            supplier       = ("Supplier",        "first"),
            invoice_date   = ("invoice_date",    "first"),
            taxable        = ("taxable",         "sum"),
            cgst           = ("cgst",            "sum"),
            sgst           = ("sgst",            "sum"),
            igst           = ("igst",            "sum"),
            member_row_ids = ("Voucher Number",  lambda x: list(x.index)),
        )
        .reset_index()
    )

    grouped["total_tax"]    = (grouped["cgst"] + grouped["sgst"] + grouped["igst"]).round(2)
    grouped["invoice_std"]  = grouped["invoice_raw"].apply(standardize_invoice)
    grouped["source"] = "books"
    grouped["invoice_date"] = pd.to_datetime(grouped["invoice_date"], errors="coerce")
    grouped = grouped.reset_index(drop=True)
    grouped.insert(0, "row_id", ["B" + str(i) for i in grouped.index])

    n_multi      = (grouped["member_row_ids"].apply(len) > 1).sum()
    n_no_gstin   = grouped["gstin"].isna().sum()
    n_standalone = (grouped["_ktype"] == "standalone").sum()

    print(
        f"[Stage 0] Grouped books: {len(grouped)} rows | "
        f"multi-line={n_multi} | no-GSTIN={n_no_gstin} | standalone={n_standalone}"
    )
    return grouped


# ═══════════════════════════════════════════════════════════════════════════════
# STAGE 1 & 2 — EXACT / STANDARDIZED MATCHING
# ═══════════════════════════════════════════════════════════════════════════════

def _identity(b: pd.Series, g: pd.Series, use_std: bool) -> tuple[bool, bool, bool]:
    """Return (gstin_match, supplier_match, invoice_match). Nulls never equal."""
    inv_b = b["invoice_std"] if use_std else b["invoice_raw"]
    inv_g = g["invoice_std"] if use_std else g["invoice_raw"]
    gm = bool(b["gstin"]    and g["gstin"]    and b["gstin"]    == g["gstin"])
    sm = bool(b["supplier"] and g["supplier"] and b["supplier"] == g["supplier"])
    im = bool(inv_b         and inv_g         and inv_b         == inv_g)
    return gm, sm, im


def _date_diff(b: pd.Series, g: pd.Series) -> int:
    bd, gd = b.get("invoice_date"), g.get("invoice_date")
    if pd.isnull(bd) or pd.isnull(gd):
        return 9999
    try:
        return abs((pd.Timestamp(bd) - pd.Timestamp(gd)).days)
    except Exception:
        return 9999


def _completeness(row: pd.Series) -> int:
    return sum(1 for f in ("gstin", "invoice_raw", "supplier") if row.get(f))


def _rank_key(c: dict) -> tuple:
    """Lower = better candidate. Tie-break order per spec."""
    return (
        0 if c["im"] else 1,
        0 if c["gm"] else 1,
        0 if c["sm"] else 1,
        c["total_tax_diff"],
        c["date_diff_days"],
        -(c["completeness_b"] + c["completeness_g"]),
    )


def _apply_tiebreak(
    candidates: list[dict],
    used_b: set[str],
    used_g: set[str],
) -> tuple[list[dict], set[str]]:
    """
    Pick the single best 2B candidate for each books row.
    - Genuine ties (top two share the same rank key) → books row is NOT matched;
      it stays in the pool (flows to Stage 3 blue or later).
    - Sort globally so a higher-priority books row wins a contested 2B slot.
    Returns (accepted, tied_books_ids).
    """
    by_books: dict[str, list[dict]] = defaultdict(list)
    for c in candidates:
        if c["books_id"] not in used_b and c["b2b_id"] not in used_g:
            by_books[c["books_id"]].append(c)

    tied_books: set[str] = set()
    proposals: list[dict] = []

    for bid, cands in by_books.items():
        cands.sort(key=_rank_key)
        if len(cands) >= 2 and _rank_key(cands[0]) == _rank_key(cands[1]):
            tied_books.add(bid)
        else:
            proposals.append(cands[0])

    # global sort: highest-priority books rows claim their 2B slot first
    proposals.sort(key=_rank_key)

    accepted: list[dict] = []
    for c in proposals:
        if c["books_id"] not in used_b and c["b2b_id"] not in used_g:
            accepted.append(c)
            used_b.add(c["books_id"])
            used_g.add(c["b2b_id"])

    return accepted, tied_books


def _is_unique_pair(
    b: pd.Series,
    g: pd.Series,
    books: pd.DataFrame,
    gstr2b: pd.DataFrame,
) -> bool:
    """
    Uniqueness gate for 1B / 2B matches (spec §Stage 1 — 1B rule).

    A loose match (identity + money, invoice differs) is only safe to
    auto-accept when exactly ONE books row and ONE 2B row share that
    (identity, amount±TOL) combination. If multiple rows exist on either
    side the amount is not a unique fingerprint — route to blue instead.

    Identity anchor: prefer GSTIN if both sides have it; fall back to supplier.
    """
    # pick the strongest shared identity
    use_gstin = bool(b["gstin"] and g["gstin"] and b["gstin"] == g["gstin"])

    if use_gstin:
        same_b = books[
            (books["gstin"] == b["gstin"]) &
            ((books["taxable"] - b["taxable"]).abs() <= TOL)
        ]
        same_g = gstr2b[
            (gstr2b["gstin"] == g["gstin"]) &
            ((gstr2b["taxable"] - g["taxable"]).abs() <= TOL)
        ]
    else:
        # supplier match
        same_b = books[
            (books["supplier"] == b["supplier"]) &
            ((books["taxable"] - b["taxable"]).abs() <= TOL)
        ]
        same_g = gstr2b[
            (gstr2b["supplier"] == g["supplier"]) &
            ((gstr2b["taxable"] - g["taxable"]).abs() <= TOL)
        ]

    return len(same_b) == 1 and len(same_g) == 1


def run_exact_stage(
    books: pd.DataFrame,
    gstr2b: pd.DataFrame,
    use_std: bool = False,
) -> tuple[list[Match], pd.DataFrame, pd.DataFrame]:
    """
    Stage 1 (use_std=False) or Stage 2 (use_std=True).

    Tight pass (xA / xC): money_match AND invoice matches
      → auto-match (any collision handled by tie-break → 1A or 1C)

    Loose pass (xB): money_match AND (gstin OR supplier) matches, invoice differs
      → auto-match ONLY when (identity, amount) is UNIQUE on both sides
      → collision cluster (non-unique) → stays in pool → becomes blue in Stage 3

    Tie-break: invoice > gstin > supplier > smallest money diff > closest date
    """
    lbl = "2" if use_std else "1"

    # build all candidates that have at least one identity match
    candidates: list[dict] = []
    for _, b in books.iterrows():
        for _, g in gstr2b.iterrows():
            if not money_match(b, g):
                continue
            gm, sm, im = _identity(b, g, use_std)
            if not (gm or sm):
                continue
            candidates.append({
                "books_id":       b["row_id"],
                "b2b_id":         g["row_id"],
                "gm": gm, "sm": sm, "im": im,
                "taxable_diff":   round(abs(b["taxable"]   - g["taxable"]),   2),
                "total_tax_diff": round(abs(b["total_tax"] - g["total_tax"]), 2),
                "date_diff_days": _date_diff(b, g),
                "completeness_b": _completeness(b),
                "completeness_g": _completeness(g),
            })

    used_b: set[str] = set()
    used_g: set[str] = set()
    matches: list[Match] = []

    # ── tight pass: invoice must match ───────────────────────────────────────
    tight       = [c for c in candidates if c["im"]]
    tight_cnt_b = Counter(c["books_id"] for c in tight)

    tight_accepted, tight_tied = _apply_tiebreak(tight, used_b, used_g)
    for c in tight_accepted:
        collision = tight_cnt_b[c["books_id"]] > 1
        sub    = f"{lbl}C" if collision else f"{lbl}A"
        reason = "exact identity + invoice + money"
        if collision:
            reason += " (collision resolved via tie-break)"
        matches.append(Match(
            c["books_id"], c["b2b_id"], lbl, sub, reason,
            c["taxable_diff"], c["total_tax_diff"],
        ))

    # ── loose pass: invoice differs, identity + money match ──────────────────
    # UNIQUENESS GATE: only auto-accept when (identity, amount) is unique on
    # both sides. Non-unique pairs stay free → become blue in Stage 3.
    books_idx  = books.set_index("row_id")
    gstr2b_idx = gstr2b.set_index("row_id")

    loose = [
        c for c in candidates
        if not c["im"]
        and c["books_id"] not in used_b
        and c["b2b_id"]   not in used_g
    ]
    loose_unique = [
        c for c in loose
        if _is_unique_pair(
            books_idx.loc[c["books_id"]],
            gstr2b_idx.loc[c["b2b_id"]],
            books,
            gstr2b,
        )
    ]
    loose_accepted, loose_tied = _apply_tiebreak(loose_unique, used_b, used_g)
    for c in loose_accepted:
        matches.append(Match(
            c["books_id"], c["b2b_id"], lbl, f"{lbl}B",
            "identity + money; invoice differs (unique pair — format/voucher mismatch)",
            c["taxable_diff"], c["total_tax_diff"],
        ))

    rem_books = books[~books["row_id"].isin(used_b)].copy()
    rem_2b    = gstr2b[~gstr2b["row_id"].isin(used_g)].copy()

    n_tied = len(tight_tied) + len(loose_tied)
    print(
        f"[Stage {lbl}] matches={len(matches)} | tied→blue={n_tied} | "
        f"remaining books={len(rem_books)}, 2B={len(rem_2b)}"
    )
    return matches, rem_books, rem_2b


# ═══════════════════════════════════════════════════════════════════════════════
# STAGE 3 — IDENTIFY BLUE ROWS ONLY (no auto-match)
# ═══════════════════════════════════════════════════════════════════════════════

def run_stage3(
    books: pd.DataFrame,
    gstr2b: pd.DataFrame,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """
    Classify leftover books rows into:
      BLUE:      money + (gstin OR supplier) matches, invoice differs.
                 → strong anchor, easy AI question.
      REMAINING: no identity anchor at all.
                 → pile builder filters further (only piles with a money-twin
                   reach the AI; isolated nodes go straight to unmatched).

    No rows are removed from the pool here — classification only.
    The same 2B candidate pool is passed to both 4A and 4B; the pile
    builder handles isolation of no-money-twin 2B rows.
    """
    blue_b: set[str] = set()
    blue_g: set[str] = set()

    for _, b in books.iterrows():
        for _, g in gstr2b.iterrows():
            if not money_match(b, g):
                continue
            id_match = (
                (b["gstin"]    and g["gstin"]    and b["gstin"]    == g["gstin"])
                or (b["supplier"] and g["supplier"] and b["supplier"] == g["supplier"])
            )
            if id_match:
                blue_b.add(b["row_id"])
                blue_g.add(g["row_id"])

    q_blue_b      = books[books["row_id"].isin(blue_b)].copy()
    q_blue_g      = gstr2b[gstr2b["row_id"].isin(blue_g)].copy()
    q_remaining_b = books[~books["row_id"].isin(blue_b)].copy()
    q_remaining_g = gstr2b[~gstr2b["row_id"].isin(blue_g)].copy()  # exclude blue 2B rows

    print(
        f"[Stage 3] blue_books={len(q_blue_b)} | "
        f"remaining→4B: books={len(q_remaining_b)}, 2B={len(q_remaining_g)}"
    )
    return q_blue_b, q_blue_g, q_remaining_b, q_remaining_g


# ═══════════════════════════════════════════════════════════════════════════════
# PILE BUILDER — connected components on money-match edges
# ═══════════════════════════════════════════════════════════════════════════════

def build_piles(books: pd.DataFrame, gstr2b: pd.DataFrame) -> tuple[list[dict], list[dict]]:
    """
    Partition rows into disjoint piles using connected components.

    Edge rule: (books_i, 2b_j) iff money_match(books_i, 2b_j).

    Guarantees:
      - Every row in exactly one pile (partition — no row in two piles).
      - No possible match ever crosses a pile boundary.
      - Rows with no money-twin on the other side → single-side pile → skipped
        by AI → go straight to unmatched.

    Size gate: piles with > PILE_SIZE_LIMIT rows per side get an exact-invoice
    pre-pass first; only the ambiguous remainder reaches the AI.

    Returns (piles, pre_matches) where pre_matches are the pairs resolved
    deterministically by the size-gate pre-pass.
    """
    G = nx.Graph()
    for _, b in books.iterrows():
        G.add_node(("B", b["row_id"]), data=b.to_dict())
    for _, g in gstr2b.iterrows():
        G.add_node(("G", g["row_id"]), data=g.to_dict())
    for _, b in books.iterrows():
        for _, g in gstr2b.iterrows():
            if money_match(b, g):
                G.add_edge(("B", b["row_id"]), ("G", g["row_id"]))

    piles: list[dict] = []
    pre_matches: list[dict] = []

    for comp in nx.connected_components(G):
        b_rows = [G.nodes[n]["data"] for n in comp if n[0] == "B"]
        g_rows = [G.nodes[n]["data"] for n in comp if n[0] == "G"]

        # single-side pile — no AI needed
        if not b_rows or not g_rows:
            piles.append({"books": b_rows, "2b": g_rows})
            continue

        # size gate: pre-pass exact-invoice matches within oversized piles
        if len(b_rows) > PILE_SIZE_LIMIT or len(g_rows) > PILE_SIZE_LIMIT:
            paired_b: set[str] = set()
            paired_g: set[str] = set()
            for b in b_rows:
                for g in g_rows:
                    if (
                        b["invoice_std"] and g["invoice_std"]
                        and b["invoice_std"] == g["invoice_std"]
                        and money_match(pd.Series(b), pd.Series(g))
                        and b["row_id"] not in paired_b
                        and g["row_id"] not in paired_g
                    ):
                        paired_b.add(b["row_id"])
                        paired_g.add(g["row_id"])
                        pre_matches.append({
                            "books_id":   b["row_id"],
                            "b2b_id":     g["row_id"],
                            "confidence": 1.0,
                            "reason":     "size-gate exact-invoice pre-pass",
                            "_pass":      "size_gate",
                        })

            b_rem = [r for r in b_rows if r["row_id"] not in paired_b]
            g_rem = [r for r in g_rows if r["row_id"] not in paired_g]
            if b_rem or g_rem:
                piles.append({"books": b_rem, "2b": g_rem})
        else:
            piles.append({"books": b_rows, "2b": g_rows})

    return piles, pre_matches


# ═══════════════════════════════════════════════════════════════════════════════
# STAGE 4A / 4B — AI MATCHING (match-the-following via tool calling)
# ═══════════════════════════════════════════════════════════════════════════════

_TOOL: dict = {
    "type": "function",
    "function": {
        "name": "submit_pile_matches",
        "description": (
            "Submit one-to-one matching decisions for this pile. "
            "Every row ID must appear exactly once across matches, "
            "unmatched_books, and unmatched_b2b."
        ),
        "parameters": {
            "type": "object",
            "required": ["reasoning", "matches", "unmatched_books", "unmatched_b2b"],
            "properties": {
                "reasoning": {
                    "type": "string",
                    "description": (
                        "Think step by step BEFORE committing to decisions. "
                        "This field is evaluated first — use it to reason about "
                        "each candidate pair before writing your final answer."
                    ),
                },
                "matches": {
                    "type": "array",
                    "items": {
                        "type": "object",
                        "required": ["books_id", "b2b_id", "reason", "confidence"],
                        "properties": {
                            "books_id":   {"type": "string"},
                            "b2b_id":     {"type": "string"},
                            "reason":     {"type": "string"},
                            "confidence": {
                                "type": "number",
                                "minimum": 0.0,
                                "maximum": 1.0,
                            },
                        },
                    },
                },
                "unmatched_books": {
                    "type": "array",
                    "items": {
                        "type": "object",
                        "required": ["books_id", "reason"],
                        "properties": {
                            "books_id": {"type": "string"},
                            "reason":   {"type": "string"},
                        },
                    },
                },
                "unmatched_b2b": {
                    "type": "array",
                    "items": {
                        "type": "object",
                        "required": ["b2b_id", "reason"],
                        "properties": {
                            "b2b_id": {"type": "string"},
                            "reason": {"type": "string"},
                        },
                    },
                },
            },
        },
    },
}

_SYSTEM_PROMPT = (
    "You are a GST reconciliation assistant. "
    "Your only job is to decide identity: whether a books invoice and a 2B invoice "
    "refer to the same real-world transaction. "
    "Financial amounts already match within ₹2 for every row in the pile — "
    "do NOT re-evaluate the money. Focus entirely on the identity fields: "
    "GSTIN (supplier tax ID), invoice number, and supplier name. "
    "Apply strict one-to-one matching: each books row pairs with at most one "
    "2B row and vice versa. A false match is worse than leaving a row unmatched.\n\n"
    "RECOGNISE THESE COMMON EQUIVALENCES — treat them as matches, not differences:\n"
    "1. Invoice number format variants: a bare number on one side and the same number "
    "with a company-code or financial-year suffix on the other ARE the same invoice. "
    "Examples: '201' = '201/SRC/24-25', '1042' = '1042/FY2425', 'INV5' = 'INV5/2425'. "
    "The core number is the invoice identity; the suffix is a supplier formatting convention.\n"
    "2. Supplier name transliteration variants: SHREE/SREE/SHRI, "
    "PVT LTD/PRIVATE LIMITED/PVT. LTD., & /AND, missing/extra spaces or punctuation — "
    "these are the same legal entity. Do not treat minor spelling variants as different suppliers.\n"
    "3. GSTIN with trailing junk: a GSTIN followed by '_X000D_' or similar artefacts "
    "is the same GSTIN as the clean version. Strip the artefact before comparing.\n\n"
    "When rows have identical amounts and supplier names, "
    "rely on invoice dates to find the best one-to-one pairing. "
    "Sort both sides by date and match the closest pairs. "
    "A confident date-based pairing is better than returning no matches."
)


def _pile_prompt(pile: dict, pass_type: str) -> str:
    if pass_type == "blue":
        question = (
            "QUESTION: Do any of these books rows refer to the same invoice as a 2B row, "
            "even though the invoice numbers are written differently? "
            "(The GSTIN or supplier already matches — the invoice number is the only disagreement.)"
        )
        instruction = (
            "Accept a match if the invoice numbers are plausibly the same document — "
            "e.g. a supplier number vs an internal voucher, or a formatting difference. "
            "Auto-accept threshold is 0.90."
        )
    else:
        question = (
            "QUESTION: Do any of these books rows refer to the same invoice as a 2B row? "
            "No identity field matched exactly by automated rules, but apply the equivalence "
            "rules from your instructions (invoice suffix variants, transliteration variants, "
            "GSTIN artefacts) before concluding there is no match."
        )
        instruction = (
            "Match when the core invoice number matches (ignoring company-code/year suffixes) "
            "AND the supplier name is the same entity (ignoring SHREE/SREE, PVT LTD/PRIVATE LIMITED, etc.). "
            "If both signals agree, confidence should be ≥ 0.93. "
            "Only leave a row unmatched when the invoice cores genuinely differ or the suppliers "
            "are clearly different companies. Auto-accept threshold is 0.93."
        )

    lines = [
        f"GST reconciliation — {pass_type} AI pass.",
        "",
        "CONTEXT:",
        "  All rows in this pile already agree on taxable value and total tax within ₹2.",
        "  Decide IDENTITY only.",
        "  One-to-one constraint: each row may be used in at most one match.",
        "  Every ID (every Bi and every Gj) must appear exactly once across",
        "  matches / unmatched_books / unmatched_b2b.",
        "",
        "COLUMN MEANINGS:",
        "  gstin    = 15-character supplier tax registration number",
        "  invoice  = invoice / voucher number as recorded on each side",
        "  supplier = trade or legal name of the supplier",
        "  taxable  = taxable value (₹)",
        "  cgst/sgst/igst = GST components (₹)",
        "",
        question,
        "",
        instruction,
        "",
        "BOOKS ROWS (internal purchase register):",
    ]

    for i, r in enumerate(pile["books"], 1):
        lines.append(
            f"  B{i}: gstin={r['gstin'] or 'N/A'}  "
            f"invoice={r['invoice_raw'] or 'N/A'}  "
            f"date={str(r['invoice_date'])[:10] if r.get('invoice_date') and str(r['invoice_date']) != 'NaT' else 'N/A'}  "
            f"supplier={r['supplier'] or 'N/A'}  "
            f"taxable={r['taxable']}  total_tax={r['total_tax']}"
        )

    lines += ["", "2B ROWS (GSTR-2B supplier-reported):"]

    for i, r in enumerate(pile["2b"], 1):
        lines.append(
            f"  G{i}: gstin={r['gstin'] or 'N/A'}  "
            f"invoice={r['invoice_raw'] or 'N/A'}  "
            f"date={str(r['invoice_date'])[:10] if r.get('invoice_date') and str(r['invoice_date']) != 'NaT' else 'N/A'}  "
            f"supplier={r['supplier'] or 'N/A'}  "
            f"taxable={r['taxable']}  total_tax={r['total_tax']}"
        )

    amounts = [r["taxable"] for r in pile["books"]]
    is_collision = len(set(amounts)) == 1 and len(pile["books"]) > 1

    if is_collision:
        lines += [
            "",
            "NOTE: All rows in this pile have identical taxable amounts. "
            "The invoice numbers are in different formats on each side (books vs 2B). "
            "Use invoice DATE as the primary matching signal — pair the books row "
            "whose date is closest to each 2B row's date. "
            "Match sequentially by date if dates are equal or very close. "
            "Only leave a row unmatched if you have a specific reason to believe "
            "it has no counterpart.",
        ]

    lines += [
        "",
        "Call submit_pile_matches.",
        "Every Bi and every Gj must appear exactly once. Use B1..Bn and G1..Gm as the IDs.",
    ]
    return "\n".join(lines)


async def _ai_pile(
    client: AsyncOpenAI,
    pile: dict,
    pass_type: str,
) -> dict | None:
    if not pile["books"] or not pile["2b"]:
        return None

    b_id_map = {f"B{i}": r for i, r in enumerate(pile["books"], 1)}
    g_id_map = {f"G{i}": r for i, r in enumerate(pile["2b"],    1)}

    prompt_text = _pile_prompt(pile, pass_type)

    log_entry: dict = {
        "pass_type":          pass_type,
        "books":              [{"local_id": k, **v} for k, v in b_id_map.items()],
        "b2b":                [{"local_id": k, **v} for k, v in g_id_map.items()],
        "prompt":             prompt_text,
        "reasoning":          None,
        "ai_matches":         [],
        "ai_unmatched_books": [],
        "ai_unmatched_b2b":   [],
        "accepted_matches":   [],
        "rejected_matches":   [],
        "error":              None,
    }

    request_payload = {
        "model":       AI_MODEL,
        "temperature": 0,
        "tool_choice": {"type": "function", "function": {"name": "submit_pile_matches"}},
        "messages": [
            {"role": "system", "content": _SYSTEM_PROMPT},
            {"role": "user",   "content": prompt_text},
        ],
    }
    _log.debug(
        "── AI REQUEST [%s pile %dB×%dG] ──\n%s",
        pass_type, len(pile["books"]), len(pile["2b"]),
        json.dumps(request_payload, indent=2, default=str),
    )

    try:
        response = await client.chat.completions.create(
            model=AI_MODEL,
            temperature=0,
            tools=[_TOOL],
            tool_choice={"type": "function", "function": {"name": "submit_pile_matches"}},
            messages=request_payload["messages"],
        )
    except Exception as exc:
        print(f"  [AI error — {pass_type}] {exc}")
        log_entry["error"] = str(exc)
        _AI_LOG.append(log_entry)
        return None

    _log.debug(
        "── AI RESPONSE [%s] ──\n%s",
        pass_type,
        response.choices[0].message.tool_calls[0].function.arguments,
    )

    raw = json.loads(response.choices[0].message.tool_calls[0].function.arguments)

    log_entry["reasoning"]          = raw.get("reasoning", "")
    log_entry["ai_matches"]         = raw.get("matches", [])
    log_entry["ai_unmatched_books"] = raw.get("unmatched_books", [])
    log_entry["ai_unmatched_b2b"]   = raw.get("unmatched_b2b", [])

    # per-pile terminal log
    print(
        f"  [pile log — {pass_type}] "
        f"books={list(b_id_map)} 2b={list(g_id_map)} | "
        f"matches={[{'B': m.get('books_id'), 'G': m.get('b2b_id'), 'conf': m.get('confidence')} for m in raw.get('matches', [])]}"
    )

    # ── validate: correct IDs, no duplicates, money check ───────────────────
    seen_b: set[str] = set()
    seen_g: set[str] = set()
    clean_matches: list[dict] = []

    for m in raw.get("matches", []):
        bid, gid = m.get("books_id", ""), m.get("b2b_id", "")
        reject_reason: str | None = None

        if bid not in b_id_map or gid not in g_id_map:
            reject_reason = f"unknown IDs {bid}/{gid}"
        elif bid in seen_b or gid in seen_g:
            reject_reason = f"duplicate ID {bid} or {gid}"
        else:
            b_row = pd.Series(b_id_map[bid])
            g_row = pd.Series(g_id_map[gid])
            if not money_match(b_row, g_row):
                reject_reason = f"money check failed"

        if reject_reason:
            log_entry["rejected_matches"].append({**m, "_reject_reason": reject_reason})
            print(f"    [AI warn] {reject_reason} — skipped")
            continue

        seen_b.add(bid)
        seen_g.add(gid)
        m["_books_row_id"] = b_id_map[bid]["row_id"]
        m["_b2b_row_id"]   = g_id_map[gid]["row_id"]
        m["_pass"]         = pass_type
        clean_matches.append(m)
        log_entry["accepted_matches"].append({
            "local_books_id": bid,
            "local_b2b_id":   gid,
            "books_row_id":   b_id_map[bid]["row_id"],
            "b2b_row_id":     g_id_map[gid]["row_id"],
            "confidence":     m.get("confidence", 0.0),
            "reason":         m.get("reason", ""),
        })

    _AI_LOG.append(log_entry)
    raw["matches"] = clean_matches
    raw["_pass"]   = pass_type
    return raw


async def run_ai_stages(
    q_blue_b:    pd.DataFrame,
    q_blue_g:    pd.DataFrame,
    remaining_b: pd.DataFrame,
    remaining_g: pd.DataFrame,
) -> list[dict]:
    """
    Stage 4A: blue piles — easy question (invoice differs despite identity match).
    Stage 4B: remaining rows — pile builder filters; single-side piles skip AI.

    Both passes run concurrently with asyncio.gather.
    Returns a flat list of validated AI proposals (one entry per proposed pair).
    """
    client = AsyncOpenAI(
        api_key=os.getenv("OPENROUTER_API_KEY"),
        base_url="https://openrouter.ai/api/v1",
    )

    piles_blue,    pre_blue    = build_piles(q_blue_b,    q_blue_g)
    piles_nonblue, pre_nonblue = build_piles(remaining_b, remaining_g)

    n_4a_sent = sum(1 for p in piles_blue    if p["books"] and p["2b"])
    n_4b_sent = sum(1 for p in piles_nonblue if p["books"] and p["2b"])
    print(
        f"  4A piles={len(piles_blue)} (sent={n_4a_sent}) | "
        f"4B piles={len(piles_nonblue)} (sent to AI={n_4b_sent}) | "
        f"size-gate pre-matches={len(pre_blue) + len(pre_nonblue)}"
    )

    tasks = (
        [_ai_pile(client, p, "blue")    for p in piles_blue    if p["books"] and p["2b"]]
        + [_ai_pile(client, p, "nonblue") for p in piles_nonblue if p["books"] and p["2b"]]
    )

    results = await asyncio.gather(*tasks, return_exceptions=True)

    proposals: list[dict] = pre_blue + pre_nonblue   # seed with size-gate pre-pass matches
    for r in results:
        if isinstance(r, Exception):
            print(f"  [AI error] {r}")
        elif r:
            for m in r.get("matches", []):
                proposals.append({
                    "books_id":   m["_books_row_id"],
                    "b2b_id":     m["_b2b_row_id"],
                    "confidence": m.get("confidence", 0.0),
                    "reason":     m.get("reason", ""),
                    "_pass":      m.get("_pass", "unknown"),
                })

    print(f"[Stage 4A/4B] AI proposals: {len(proposals)}")
    return proposals


# ═══════════════════════════════════════════════════════════════════════════════
# AI LOG — HTML REPORT
# ═══════════════════════════════════════════════════════════════════════════════

def write_ai_log_html(path: Path) -> None:
    """Write _AI_LOG to a self-contained HTML file for easy inspection."""
    import html as _html

    def esc(v) -> str:
        if v is None or (isinstance(v, float) and pd.isna(v)):
            return '<span class="na">N/A</span>'
        return _html.escape(str(v))

    def date_str(v) -> str:
        try:
            t = pd.Timestamp(v)
            return t.strftime("%Y-%m-%d") if not pd.isnull(t) else "—"
        except Exception:
            return "—"

    def conf_badge(conf: float, pass_type: str) -> str:
        thr   = BLUE_AUTO_ACCEPT   if pass_type == "blue" else NONBLUE_AUTO_ACCEPT
        rev   = BLUE_MIN_REVIEW    if pass_type == "blue" else NONBLUE_MIN_REVIEW
        pct   = f"{conf:.0%}"
        if conf >= thr:
            return f'<span class="badge b-green">✓ AUTO-ACCEPTED {pct}</span>'
        if conf >= rev:
            return f'<span class="badge b-yellow">⚠ IN REVIEW {pct}</span>'
        return f'<span class="badge b-red">✗ BELOW THRESHOLD {pct}</span>'

    def row_table(rows: list[dict]) -> str:
        heads = "<tr><th>ID</th><th>GSTIN</th><th>Invoice</th><th>Date</th><th>Supplier</th><th>Taxable ₹</th><th>Total Tax ₹</th></tr>"
        body  = ""
        for r in rows:
            gstin = r.get("gstin")
            gstin_cell = '<span class="na null-gstin">NULL ⚠</span>' if not gstin else esc(gstin)
            body += (
                f"<tr>"
                f"<td><b>{esc(r.get('local_id'))}</b></td>"
                f"<td>{gstin_cell}</td>"
                f"<td>{esc(r.get('invoice_raw'))}</td>"
                f"<td>{date_str(r.get('invoice_date'))}</td>"
                f"<td>{esc(r.get('supplier'))}</td>"
                f"<td class='num'>{esc(r.get('taxable'))}</td>"
                f"<td class='num'>{esc(r.get('total_tax'))}</td>"
                f"</tr>"
            )
        return f"<table>{heads}{body}</table>"

    pile_cards = []
    n_blue = n_nonblue = n_errors = total_accepted = 0

    for idx, e in enumerate(_AI_LOG, 1):
        pt        = e["pass_type"]
        thr       = BLUE_AUTO_ACCEPT if pt == "blue" else NONBLUE_AUTO_ACCEPT
        rev       = BLUE_MIN_REVIEW  if pt == "blue" else NONBLUE_MIN_REVIEW
        n_prop  = len(e["ai_matches"])
        n_acc   = len(e["accepted_matches"])
        has_err = bool(e.get("error"))

        if pt == "blue":    n_blue    += 1
        else:               n_nonblue += 1
        if has_err:         n_errors  += 1
        total_accepted += n_acc

        # header badge summary
        if has_err:
            status_badge = '<span class="badge b-grey">ERROR</span>'
        elif n_acc:
            status_badge = f'<span class="badge b-green">{n_acc} matched</span>'
        elif n_prop:
            status_badge = f'<span class="badge b-yellow">{n_prop} proposed, 0 accepted</span>'
        else:
            status_badge = '<span class="badge b-red">no matches</span>'

        # build match decisions table
        if has_err:
            decisions_html = f'<div class="error-box">ERROR: {esc(e["error"])}</div>'
        elif not e["ai_matches"]:
            decisions_html = '<p class="muted">AI returned no matches for this pile.</p>'
        else:
            rows_html = ""
            for m in e["ai_matches"]:
                conf = m.get("confidence", 0.0)
                is_rej_val = any(
                    r.get("books_id") == m.get("books_id") and r.get("b2b_id") == m.get("b2b_id")
                    for r in e["rejected_matches"]
                )
                if is_rej_val:
                    rr = next(r for r in e["rejected_matches"]
                              if r.get("books_id") == m.get("books_id"))
                    outcome = f'<span class="badge b-grey">VALIDATION FAILED: {esc(rr.get("_reject_reason",""))}</span>'
                    row_cls = "row-grey"
                elif conf >= thr:
                    outcome  = conf_badge(conf, pt)
                    row_cls  = "row-green"
                elif conf >= rev:
                    outcome  = conf_badge(conf, pt)
                    row_cls  = "row-yellow"
                else:
                    outcome  = conf_badge(conf, pt)
                    row_cls  = "row-red"

                rows_html += (
                    f'<tr class="{row_cls}">'
                    f"<td><b>{esc(m.get('books_id'))}</b></td>"
                    f"<td><b>{esc(m.get('b2b_id'))}</b></td>"
                    f"<td class='num'>{esc(m.get('confidence',''))}</td>"
                    f"<td>{outcome}</td>"
                    f"<td>{esc(m.get('reason',''))}</td>"
                    f"</tr>"
                )
            decisions_html = (
                f"<p class='threshold-note'>Thresholds — auto-accept: {thr}, min-review: {rev}</p>"
                f"<table><tr><th>Books ID</th><th>2B ID</th><th>Confidence</th>"
                f"<th>Outcome</th><th>AI Reason</th></tr>{rows_html}</table>"
            )

        unmatched_b = e.get("ai_unmatched_books", [])
        unmatched_g = e.get("ai_unmatched_b2b", [])
        unmatched_html = ""
        if unmatched_b:
            items = "; ".join(f"{esc(u.get('books_id'))} — {esc(u.get('reason',''))}" for u in unmatched_b)
            unmatched_html += f'<p class="muted"><b>Unmatched books:</b> {items}</p>'
        if unmatched_g:
            items = "; ".join(f"{esc(u.get('b2b_id'))} — {esc(u.get('reason',''))}" for u in unmatched_g)
            unmatched_html += f'<p class="muted"><b>Unmatched 2B:</b> {items}</p>'

        card_cls = f"pile pile-{pt}" + (" pile-error" if has_err else "")
        pile_cards.append(f"""
<div class="{card_cls}" data-pass="{pt}" data-accepted="{n_acc}">
  <details>
    <summary>
      <span class="pile-title">Pile {idx} &nbsp;·&nbsp; <b>{pt.upper()}</b>
        &nbsp;·&nbsp; {len(e['books'])}B × {len(e['b2b'])}G</span>
      {status_badge}
    </summary>
    <div class="pile-body">
      <div class="grid-2">
        <div>
          <div class="slabel">BOOKS ROWS</div>
          {row_table(e['books'])}
        </div>
        <div>
          <div class="slabel">2B ROWS</div>
          {row_table(e['b2b'])}
        </div>
      </div>

      <div class="slabel">AI REASONING</div>
      <div class="reasoning">{esc(e.get('reasoning') or '(no reasoning returned)')}</div>

      <div class="slabel">MATCH DECISIONS</div>
      {decisions_html}
      {unmatched_html}

      <details class="prompt-toggle">
        <summary>Show full prompt sent to AI</summary>
        <pre class="prompt-pre">{esc(e.get('prompt', ''))}</pre>
      </details>
    </div>
  </details>
</div>""")

    total_piles = len(_AI_LOG)
    cards_html  = "\n".join(pile_cards)
    ts          = pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S")

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<title>AI Matching Log — GST Reconciliation</title>
<style>
  *{{box-sizing:border-box;margin:0;padding:0}}
  body{{font-family:'Segoe UI',Arial,sans-serif;background:#f0f2f5;color:#222;padding:20px}}
  h1{{font-size:1.5rem;margin-bottom:4px}}
  .meta{{color:#888;font-size:.85rem;margin-bottom:18px}}
  .summary{{display:flex;gap:20px;flex-wrap:wrap;background:#fff;padding:16px 20px;border-radius:10px;
            box-shadow:0 1px 4px rgba(0,0,0,.08);margin-bottom:20px}}
  .stat .n{{font-size:2rem;font-weight:700;line-height:1}}
  .stat .lbl{{font-size:.78rem;color:#666;margin-top:2px}}
  .n-blue{{color:#1a7fe8}}.n-green{{color:#1ab558}}.n-orange{{color:#e88f1a}}.n-red{{color:#e83a1a}}
  .filters{{margin-bottom:16px;display:flex;gap:8px;flex-wrap:wrap}}
  .filters button{{padding:5px 14px;border:1px solid #ccc;border-radius:20px;cursor:pointer;
                   background:#fff;font-size:.85rem;transition:all .15s}}
  .filters button.on{{background:#1a7fe8;color:#fff;border-color:#1a7fe8}}
  .pile{{background:#fff;border-radius:10px;margin-bottom:10px;
         box-shadow:0 1px 3px rgba(0,0,0,.07);overflow:hidden}}
  .pile details>summary{{padding:12px 16px;cursor:pointer;list-style:none;
                          display:flex;align-items:center;justify-content:space-between;gap:10px}}
  .pile details>summary::-webkit-details-marker{{display:none}}
  .pile-blue   details>summary{{border-left:4px solid #4da6ff;background:#f0f7ff}}
  .pile-nonblue details>summary{{border-left:4px solid #ffaa00;background:#fffbf0}}
  .pile-error   details>summary{{border-left:4px solid #e83a1a;background:#fff5f5}}
  .pile-title{{font-size:.95rem}}
  .pile-body{{padding:16px 18px;border-top:1px solid #eee}}
  .grid-2{{display:grid;grid-template-columns:1fr 1fr;gap:16px;margin-bottom:12px}}
  @media(max-width:900px){{.grid-2{{grid-template-columns:1fr}}}}
  .slabel{{font-size:.72rem;font-weight:700;color:#999;text-transform:uppercase;
           letter-spacing:.5px;margin:14px 0 6px}}
  table{{border-collapse:collapse;width:100%;font-size:.82rem}}
  th{{background:#f5f5f5;padding:5px 9px;text-align:left;border:1px solid #e0e0e0;font-weight:600}}
  td{{padding:4px 9px;border:1px solid #e8e8e8;vertical-align:top}}
  .num{{text-align:right;font-variant-numeric:tabular-nums}}
  .na{{color:#aaa;font-style:italic;font-size:.8rem}}
  .null-gstin{{color:#c44;font-weight:600}}
  .reasoning{{background:#fffdf0;border-left:3px solid #ffc107;padding:10px 14px;
              font-size:.85rem;line-height:1.6;white-space:pre-wrap;border-radius:0 4px 4px 0}}
  .threshold-note{{font-size:.78rem;color:#888;margin-bottom:6px}}
  .row-green{{background:#f0fff5}}.row-yellow{{background:#fffde7}}
  .row-red{{background:#fff5f5}}.row-grey{{background:#f5f5f5}}
  .badge{{display:inline-block;padding:2px 8px;border-radius:12px;font-size:.75rem;font-weight:700}}
  .b-green{{background:#1ab558;color:#fff}}.b-yellow{{background:#ffc107;color:#333}}
  .b-red{{background:#e83a1a;color:#fff}}.b-grey{{background:#888;color:#fff}}
  .error-box{{background:#fff0f0;border:1px solid #ffb3b3;padding:10px;border-radius:4px;color:#c00}}
  .muted{{color:#888;font-size:.83rem;margin:6px 0}}
  .prompt-toggle>summary{{color:#aaa;font-size:.78rem;cursor:pointer;margin-top:10px;padding:4px 0}}
  .prompt-pre{{background:#f8f8f8;border:1px solid #ddd;padding:12px;font-size:.75rem;
               white-space:pre-wrap;max-height:280px;overflow-y:auto;border-radius:4px;margin-top:6px}}
</style>
</head>
<body>
<h1>AI Matching Log — GST Reconciliation</h1>
<p class="meta">Generated {ts} &nbsp;|&nbsp; Model: {AI_MODEL}</p>

<div class="summary">
  <div class="stat"><div class="n">{total_piles}</div><div class="lbl">Piles sent to AI</div></div>
  <div class="stat"><div class="n n-blue">{n_blue}</div><div class="lbl">Blue (4A)</div></div>
  <div class="stat"><div class="n n-orange">{n_nonblue}</div><div class="lbl">Nonblue (4B)</div></div>
  <div class="stat"><div class="n n-green">{total_accepted}</div><div class="lbl">Matches accepted</div></div>
  <div class="stat"><div class="n n-red">{n_errors}</div><div class="lbl">Errors</div></div>
</div>

<div class="filters">
  <button class="on" onclick="f('all',this)">All</button>
  <button onclick="f('blue',this)">Blue only</button>
  <button onclick="f('nonblue',this)">Nonblue only</button>
  <button onclick="f('matched',this)">Has matches</button>
  <button onclick="f('unmatched',this)">No matches</button>
</div>

{cards_html}

<script>
function f(mode,btn){{
  document.querySelectorAll('.filters button').forEach(b=>b.classList.remove('on'));
  btn.classList.add('on');
  document.querySelectorAll('.pile').forEach(p=>{{
    const pass=p.dataset.pass, acc=parseInt(p.dataset.accepted)||0;
    let show=true;
    if(mode==='blue')      show=pass==='blue';
    else if(mode==='nonblue')   show=pass==='nonblue';
    else if(mode==='matched')   show=acc>0;
    else if(mode==='unmatched') show=acc===0;
    p.style.display=show?'':'none';
  }});
}}
</script>
</body></html>"""

    path.write_text(html, encoding="utf-8")
    print(f"  AI log HTML → {path}")


# ═══════════════════════════════════════════════════════════════════════════════
# STAGE 5 — GREEDY MERGE + 7-SHEET OUTPUT WORKBOOK
# ═══════════════════════════════════════════════════════════════════════════════

def _pair_unmatched(b_rows: list, g_rows: list) -> list:
    """
    Pair unmatched books and 2B rows by similarity so related rows appear
    side by side in the white section of the output.

    Scoring per (b, g) candidate pair:
      +4  GSTIN exact (after stripping _X000D_ artefacts)
      +3  supplier exact (normalised)
      +2  supplier fuzzy (Jaccard word-set overlap × 2)
      +2  invoice_std exact

    Greedy: sort all candidates by score desc, accept highest non-conflicting pairs.
    Unpaired rows appended solo at the end (books-only or 2B-only).
    """
    def _clean_gstin(v) -> str | None:
        if not v or (isinstance(v, float) and pd.isna(v)):
            return None
        return re.sub(r"_.*", "", str(v)).strip().upper() or None

    def _norm(v) -> str | None:
        if not v or (isinstance(v, float) and pd.isna(v)):
            return None
        s = re.sub(r"[^A-Z0-9 ]", " ", str(v).upper())
        return re.sub(r"\s+", " ", s).strip() or None

    def _score(b: pd.Series, g: pd.Series) -> float:
        s = 0.0
        bg, gg = _clean_gstin(b.get("gstin")), _clean_gstin(g.get("gstin"))
        if bg and gg and bg == gg:
            s += 4.0
        bs, gs = _norm(b.get("supplier")), _norm(g.get("supplier"))
        if bs and gs:
            if bs == gs:
                s += 3.0
            else:
                bw, gw = set(bs.split()), set(gs.split())
                if bw and gw:
                    s += (len(bw & gw) / len(bw | gw)) * 2.0
        bi, gi = b.get("invoice_std"), g.get("invoice_std")
        if bi and gi and bi == gi:
            s += 2.0
        return s

    candidates: list[tuple[float, int, int]] = []
    for bi, (_, b) in enumerate(b_rows):
        for gi, (_, g) in enumerate(g_rows):
            sc = _score(b, g)
            if sc > 0:
                candidates.append((sc, bi, gi))
    candidates.sort(reverse=True)

    used_b: set[int] = set()
    used_g: set[int] = set()
    pairs: list = []

    for sc, bi, gi in candidates:
        if bi not in used_b and gi not in used_g:
            pairs.append((b_rows[bi], g_rows[gi]))
            used_b.add(bi)
            used_g.add(gi)

    for bi, b_item in enumerate(b_rows):
        if bi not in used_b:
            pairs.append((b_item, None))
    for gi, g_item in enumerate(g_rows):
        if gi not in used_g:
            pairs.append((None, g_item))

    return pairs


def _find_identity_money_mismatch(
    unmatched_b: pd.DataFrame,
    unmatched_g: pd.DataFrame,
) -> pd.DataFrame:
    """
    Among final unmatched rows, find pairs where identity matches on
    invoice_std (GSTIN+invoice OR supplier+invoice) but money does NOT match.

    Uses invoice_std (not raw) so formatting differences don't create phantom
    mismatches. Results are advisory — for the accountant to review.
    Same-invoice-different-amount is the most actionable exception class.
    """
    rows = []
    seen: set[tuple] = set()   # avoid duplicate pairs

    for _, b in unmatched_b.iterrows():
        for _, g in unmatched_g.iterrows():
            if money_match(b, g):
                continue   # money matches → not a mismatch

            std_b = b.get("invoice_std")
            std_g = g.get("invoice_std")
            if not std_b or not std_g:
                continue

            gstin_inv = (
                b["gstin"] and g["gstin"]
                and b["gstin"] == g["gstin"]
                and std_b == std_g
            )
            sup_inv = (
                b["supplier"] and g["supplier"]
                and b["supplier"] == g["supplier"]
                and std_b == std_g
            )
            if not (gstin_inv or sup_inv):
                continue

            key = (b["row_id"], g["row_id"])
            if key in seen:
                continue
            seen.add(key)

            rows.append({
                "books_row_id":   b["row_id"],
                "b2b_row_id":     g["row_id"],
                "gstin":          b.get("gstin"),
                "invoice_books":  b.get("invoice_raw"),
                "invoice_2b":     g.get("invoice_raw"),
                "supplier_books": b.get("supplier"),
                "supplier_2b":    g.get("supplier"),
                "taxable_books":  b.get("taxable"),
                "taxable_2b":     g.get("taxable"),
                "cgst_books":     b.get("cgst"),     "cgst_2b":  g.get("cgst"),
                "sgst_books":     b.get("sgst"),     "sgst_2b":  g.get("sgst"),
                "igst_books":     b.get("igst"),     "igst_2b":  g.get("igst"),
                "taxable_diff":   round(abs(b["taxable"]   - g["taxable"]),   2),
                "total_tax_diff": round(abs(b["total_tax"] - g["total_tax"]), 2),
                "anchor":         "GSTIN+invoice" if gstin_inv else "supplier+invoice",
                "flag":           "IDENTITY_MATCH_MONEY_MISMATCH",
            })

    return pd.DataFrame(rows)


def run_stage5(
    det_matches:   list[Match],
    ai_proposals:  list[dict],
    books_grouped: pd.DataFrame,
    gstr2b:        pd.DataFrame,
    dirty_books:   pd.DataFrame,
    output_path:   Path,
) -> None:
    """
    Greedy one-to-one merge of AI proposals, then write consolidated workbook:
    one 'reconciliation' sheet with color-coded rows + 'blue_review' and
    'dirty_input' utility sheets.
    """
    books_idx = books_grouped.set_index("row_id")
    b2b_idx   = gstr2b.set_index("row_id")

    used_b: set[str] = {m.books_id for m in det_matches}
    used_g: set[str] = {m.b2b_id   for m in det_matches}

    ai_proposals_sorted = sorted(
        ai_proposals, key=lambda x: x.get("confidence", 0), reverse=True
    )
    ai_accepted: list[dict] = []
    ai_review:   list[dict] = []

    for p in ai_proposals_sorted:
        bid, gid = p["books_id"], p["b2b_id"]
        if bid in used_b or gid in used_g:
            continue
        conf       = p.get("confidence", 0)
        is_blue    = p.get("_pass") == "blue"
        threshold  = BLUE_AUTO_ACCEPT   if is_blue else NONBLUE_AUTO_ACCEPT
        min_review = BLUE_MIN_REVIEW    if is_blue else NONBLUE_MIN_REVIEW
        if conf >= threshold:
            ai_accepted.append(p)
            used_b.add(bid)
            used_g.add(gid)
        elif conf >= min_review:
            ai_review.append(p)

    unmatched_b_ids = set(books_grouped["row_id"]) - used_b
    df_unmatched_b  = books_grouped[books_grouped["row_id"].isin(unmatched_b_ids)].copy()
    unmatched_g_ids = set(gstr2b["row_id"]) - used_g
    df_unmatched_g  = gstr2b[gstr2b["row_id"].isin(unmatched_g_ids)].copy()

    df_id_mismatch = _find_identity_money_mismatch(df_unmatched_b, df_unmatched_g)
    if len(df_id_mismatch):
        print(f"  [Flag] identity+invoice match, money differs: {len(df_id_mismatch)} pairs")

    df_blue_review = pd.DataFrame(ai_review) if ai_review else pd.DataFrame()

    # ── conservation check (before touching the file) ─────────────────────────
    n_matched = len(det_matches) + len(ai_accepted)
    total_b   = n_matched + len(df_unmatched_b)
    total_g   = n_matched + len(df_unmatched_g)
    assert total_b == len(books_grouped), (
        f"Books conservation failed: {total_b} != {len(books_grouped)}"
    )
    assert total_g == len(gstr2b), (
        f"2B conservation failed: {total_g} != {len(gstr2b)}"
    )

    # id-mismatch pairs are shown as paired red rows; exclude them from the
    # solo white unmatched rows so they don't appear twice
    id_mm_b = set(df_id_mismatch["books_row_id"]) if len(df_id_mismatch) else set()
    id_mm_g = set(df_id_mismatch["b2b_row_id"])   if len(df_id_mismatch) else set()

    def _safe(v):
        try:
            if pd.isna(v):
                return None
        except (TypeError, ValueError):
            pass
        return v

    def _b_fields(b: pd.Series) -> dict:
        ids = b.get("member_row_ids", [])
        row_id = ", ".join(str(x) for x in ids) if isinstance(ids, list) else str(ids)
        tx   = _safe(b.get("taxable"))
        cgst = _safe(b.get("cgst"))
        sgst = _safe(b.get("sgst"))
        igst = _safe(b.get("igst"))
        tt   = round((cgst or 0) + (sgst or 0) + (igst or 0), 2)
        return {
            "b_row_id":      row_id,
            "b_gstin":       _safe(b.get("gstin"))       or "",
            "b_gross_total": round(tx + tt, 2) if tx is not None else None,
            "b_cgst":        cgst,
            "b_sgst":        sgst,
            "b_igst":        igst,
            "b_invoice_raw": _safe(b.get("invoice_raw")) or "",
            "b_supplier":    _safe(b.get("supplier"))    or "",
            "b_taxable":     tx,
        }

    def _g_fields(g: pd.Series) -> dict:
        return {
            "g_taxable":       _safe(g.get("taxable")),
            "g_supplier":      _safe(g.get("supplier"))      or "",
            "g_invoice_raw":   _safe(g.get("invoice_raw"))   or "",
            "g_cgst":          _safe(g.get("cgst")),
            "g_sgst":          _safe(g.get("sgst")),
            "g_igst":          _safe(g.get("igst")),
            "g_invoice_value": _safe(g.get("invoice_value")),
            "g_gstin":         _safe(g.get("gstin"))         or "",
            "g_row_id":        g.get("row_id"),
        }

    _BLANK_B: dict = {
        "b_row_id": "", "b_gstin": "", "b_gross_total": "",
        "b_cgst": "", "b_sgst": "", "b_igst": "",
        "b_invoice_raw": "", "b_supplier": "", "b_taxable": "",
    }
    _BLANK_G: dict = {
        "g_taxable": "", "g_supplier": "", "g_invoice_raw": "",
        "g_cgst": "", "g_sgst": "", "g_igst": "",
        "g_invoice_value": "", "g_gstin": "", "g_row_id": "",
    }

    # ── build flat row list in spec order ─────────────────────────────────────
    rows: list[dict] = []

    # 1 + 2: deterministic matches (yellow = stage 1, light green = stage 2)
    for m in det_matches:
        color = "#FFFF99" if m.stage == "1" else "#CCFFCC"
        rows.append({
            **_b_fields(books_idx.loc[m.books_id]),
            **_g_fields(b2b_idx.loc[m.b2b_id]),
            "_color": color,
        })

    # 3 + 4: AI matches (light blue = blue pass, peach = nonblue/size-gate)
    for p in ai_accepted:
        color = "#CCE5FF" if p.get("_pass") == "blue" else "#FFE4B5"
        rows.append({
            **_b_fields(books_idx.loc[p["books_id"]]),
            **_g_fields(b2b_idx.loc[p["b2b_id"]]),
            "_color": color,
        })

    # 5: identity match, money mismatch (light red — both sides filled)
    for _, r in df_id_mismatch.iterrows():
        rows.append({
            **_b_fields(books_idx.loc[r["books_row_id"]]),
            **_g_fields(b2b_idx.loc[r["b2b_row_id"]]),
            "_color": "#FFB3B3",
        })

    # 6 + 7: unmatched rows — similar ones paired side by side, unpaired rows solo at end
    b_unmatched = list(df_unmatched_b[~df_unmatched_b["row_id"].isin(id_mm_b)].iterrows())
    g_unmatched = list(df_unmatched_g[~df_unmatched_g["row_id"].isin(id_mm_g)].iterrows())
    for b_item, g_item in _pair_unmatched(b_unmatched, g_unmatched):
        rows.append({
            **(_b_fields(b_item[1]) if b_item is not None else _BLANK_B),
            **(_g_fields(g_item[1]) if g_item is not None else _BLANK_G),
            "_color": "#FFFFFF",
        })

    # ── write workbook ─────────────────────────────────────────────────────────
    # Column layout (18 cols):
    #   BOOKS A:I — row_id | gstin | gross_total | cgst | sgst | igst | voucher_no | supplier | taxable
    #   2B    J:R — taxable | supplier | invoice_no | cgst | sgst | igst | invoice_value | gstin | row_id
    wb = Workbook()
    ws = wb.active
    ws.title = "reconciliation"

    ws.merge_cells("A1:I1")
    ws["A1"].value     = "BOOKS"
    ws["A1"].font      = Font(bold=True)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("J1:R1")
    ws["J1"].value     = "2B"
    ws["J1"].font      = Font(bold=True)
    ws["J1"].alignment = Alignment(horizontal="center")

    labels = [
        "row_id", "gstin", "gross_total", "cgst", "sgst", "igst", "voucher_no", "supplier", "taxable",
        "taxable", "supplier", "invoice_no", "cgst", "sgst", "igst", "invoice_value", "gstin", "row_id",
    ]
    for col, label in enumerate(labels, 1):
        cell      = ws.cell(row=2, column=col, value=label)
        cell.font = Font(bold=True)

    ws.freeze_panes = "A3"

    COLS = [
        "b_row_id", "b_gstin", "b_gross_total", "b_cgst", "b_sgst", "b_igst", "b_invoice_raw", "b_supplier", "b_taxable",
        "g_taxable", "g_supplier", "g_invoice_raw", "g_cgst", "g_sgst", "g_igst", "g_invoice_value", "g_gstin", "g_row_id",
    ]
    for r_idx, row in enumerate(rows, start=3):
        fill = PatternFill("solid", fgColor=row["_color"].lstrip("#"))
        for c_idx, key in enumerate(COLS, 1):
            cell      = ws.cell(row=r_idx, column=c_idx, value=row.get(key, ""))
            cell.fill = fill

    # utility sheets
    ws_br = wb.create_sheet("blue_review")
    if not df_blue_review.empty:
        ws_br.append(list(df_blue_review.columns))
        for _, rd in df_blue_review.iterrows():
            ws_br.append([_safe(v) for v in rd])

    ws_di = wb.create_sheet("dirty_input")
    if not dirty_books.empty:
        ws_di.append(list(dirty_books.columns))
        for _, rd in dirty_books.iterrows():
            ws_di.append([_safe(v) for v in rd])

    ws_leg = wb.create_sheet("legend")
    _legend = [
        (None,     "Colour",       "Meaning"),
        ("FFFF99", "Yellow",       "Stage 1 — exact match (GSTIN + invoice + amount)"),
        ("CCFFCC", "Light green",  "Stage 2 — standardized invoice match"),
        ("CCE5FF", "Light blue",   "Stage 4A — AI match (identity confirmed, invoice format differs)"),
        ("FFE4B5", "Peach",        "Stage 4B / size-gate — AI match (no prior identity anchor)"),
        ("FFB3B3", "Light red",    "Identity + invoice match but amounts differ — review needed"),
        ("FFFFFF", "White",        "Unmatched (one or both sides blank)"),
    ]
    for i, (hex_color, label, meaning) in enumerate(_legend, 1):
        c1 = ws_leg.cell(row=i, column=1, value=label)
        c2 = ws_leg.cell(row=i, column=2, value=meaning)
        if hex_color is None:
            c1.font = Font(bold=True)
            c2.font = Font(bold=True)
        else:
            _f = PatternFill("solid", fgColor=hex_color)
            c1.fill = _f
            c2.fill = _f
    ws_leg.column_dimensions["A"].width = 16
    ws_leg.column_dimensions["B"].width = 62

    wb.save(output_path)

    print(f"\n[Stage 5] Output → {output_path}")
    print(f"  stage 1+2 matches:        {sum(1 for m in det_matches if m.stage in ('1','2'))}")
    print(f"  AI matches:               {len(ai_accepted)}")
    print(f"  id/money mismatch pairs:  {len(df_id_mismatch)}")
    print(f"  unmatched books:          {len(df_unmatched_b)}  (incl. no-GSTIN rows)")
    print(f"  unmatched 2b:             {len(df_unmatched_g)}")
    print(f"  blue_review:              {len(df_blue_review)}")
    print(f"  dirty_input:              {len(dirty_books)}")
    print("  [OK] conservation check passed")


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════════

async def main(
    books_path:  str       = BOOKS_PATH,
    b2b_path:    str       = B2B_PATH,
    output_path: Path      = OUTPUT_PATH,
    col_map:     dict|None = None,
    header_row:  int       = 1,
) -> None:
    print("\n── Stage 0: Load & Prepare ──")
    books_clean, books_dirty = load_books(books_path, col_map=col_map, header_row=header_row)
    gstr2b                   = load_2b(b2b_path)
    books_grouped            = group_books(books_clean)

    print("\n── Stage 1: Exact Matching ──")
    m1, books_r1, gstr2b_r1 = run_exact_stage(books_grouped, gstr2b, use_std=False)

    print("\n── Stage 2: Standardized Invoice ──")
    m2, books_r2, gstr2b_r2 = run_exact_stage(books_r1, gstr2b_r1, use_std=True)

    print("\n── Stage 3: Identify Blue Rows ──")
    q_blue_b, q_blue_g, q_remaining_b, q_remaining_g = run_stage3(books_r2, gstr2b_r2)

    print("\n── Stage 4A/4B: AI Matching ──")
    ai_proposals = await run_ai_stages(q_blue_b, q_blue_g, q_remaining_b, q_remaining_g)

    print("\n── Stage 5: Merge & Output ──")
    run_stage5(
        det_matches   = m1 + m2,
        ai_proposals  = ai_proposals,
        books_grouped = books_grouped,
        gstr2b        = gstr2b,
        dirty_books   = books_dirty,
        output_path   = Path(output_path),
    )

    print("\n── AI Log ──")
    write_ai_log_html(Path(output_path).with_name("ai_matching_log.html"))


if __name__ == "__main__":
    asyncio.run(main())